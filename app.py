# app_controller.py

from flask import Flask, render_template, request, redirect, url_for, flash, get_flashed_messages, render_template_string, send_file
import mysql.connector
import os
from io import BytesIO
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from openpyxl import Workbook, styles
import tempfile
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from sqlalchemy import create_engine
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.graph_objects as go


from models.db_connection import get_db_connection
from controllers.regresion_lineal import RegresionLinealModel
from controllers.analizar_correlacion import analizar_correlacion
from flask_session import Session

# Crear una instancia de la aplicación Flask
app = Flask(__name__, template_folder='views')
app.secret_key = os.urandom(24)

app.config['SESSION_TYPE'] = 'filesystem'
Session(app)

app.debug = False

# Inicializar la clase RegresionLinealModel
regresion_model = RegresionLinealModel()

#----------------------------SELECCION----------------------------

# Función para conectar a la base de datos MySQL
def connect_to_db():
    try:
        db = get_db_connection()
        if db is not None:
            return db
        else:
            return None
    except mysql.connector.Error as err:
        return None

# Función para obtener la lista de tablas en la base de datos OK
def get_table_list():
    db = connect_to_db()
    if db is not None:
        cursor = db.cursor()
        cursor.execute("SHOW TABLES")
        tables = [table[0] for table in cursor]
        table_list = []
        for table_name in tables:
            columns = get_table_columns(table_name)
            table_list.append({'name': table_name, 'columns': columns})
        db.close()
        return table_list
    else:
        return []

# Función para obtener las columnas de una tabla y sus registros
def get_table_info(table_name):
    columns = get_table_columns(table_name)
    records = get_table_records(table_name)
    return columns, records

# Función para obtener las columnas de una tabla
def get_table_columns(table_name):
    db = connect_to_db()
    if db is not None:
        cursor = db.cursor()
        cursor.execute(f"SHOW COLUMNS FROM {table_name}")
        columns = [col[0] for col in cursor.fetchall()]
        db.close()
        return columns
    else:
        return []

# Función para obtener los registros de una tabla
def get_table_records(table_name):
    db = connect_to_db()
    if db is not None:
        cursor = db.cursor()
        cursor.execute(f"SELECT * FROM {table_name}")
        records = cursor.fetchall()
        db.close()
        return records
    else:
        return []

# Función para obtener los nombres de las columnas de una tabla
def obtener_nombres_columnas(selected_table):
    try:
        db = get_db_connection()
        if db.is_connected():
            cursor = db.cursor()
            query = f"SHOW COLUMNS FROM {selected_table}"
            cursor.execute(query)
            result = cursor.fetchall()
            db.close()

            if result:
                column_names = [row[0] for row in result]
                return column_names
            else:
                raise Exception(f"No se encontraron columnas en la tabla '{selected_table}'.")
        else:
            raise Exception("La conexión a la base de datos no está activa.")
    except Exception as e:
        raise Exception(f"Error al conectar a la base de datos: {e}")

@app.route('/', methods=['GET'])
def inicio_marca():
    return render_template('inicio_marca.html')


@app.route('/index', methods=['GET'])
def index():
    tables = get_table_list()
    return render_template('seleccion.html', action='seleccion', tables=tables, selected_tables=[])



# Ruta para procesar la selección de tabla(s) en la base de datos y mostrar opciones de acción
@app.route('/seleccionar-tabla/<action>', methods=['GET', 'POST'])
def seleccionar_tabla(action):
    error_message = None
    selected_tables = []
    records = {}
    columns = {}
    selected_variables = {'x_variable': None, 'y_variable': None}  # Inicialización

    # Obtener la lista de tablas en la base de datos
    tables = get_table_list()

    if request.method == 'POST':
        selected_tables = request.form.getlist('table')
        if selected_tables:
            # Obtener las columnas y registros de todas las tablas seleccionadas
            for table_name in selected_tables:
                table_columns, table_records = get_table_info(table_name)
                columns[table_name] = table_columns
                records[table_name] = table_records

    # Obtener las variables seleccionadas si ya se han seleccionado en la primera acción
    if 'selected_variables' in request.form:
        selected_variables['x_variable'] = request.form.get('x_variable')
        selected_variables['y_variable'] = request.form.get('y_variable')

    # Pasa selected_variables al renderizar la plantilla
    return render_template('seleccion.html', action=action, tables=tables, selected_tables=selected_tables,
                           error_message=error_message, columns=columns, records=records,
                           selected_variables=selected_variables)

#----------------------------CORRELACION----------------------------

# Ruta que permite al usuario seleccionar las variables para el análisis de correlación en una tabla de la base de datos.
@app.route('/seleccionar-variables-correlacion/<table_name>', methods=['GET', 'POST'])
def seleccionar_variables_correlacion(table_name):
    error_message = None
    selected_columns = []

    # Obtener la lista de columnas y registros de la tabla seleccionada
    columns, records = get_table_info(table_name)

    if request.method == 'POST':
        selected_columns = request.form.getlist('variables[]')  

    # Renderizar la plantilla 'seleccion.html' con los datos obtenidos
    return render_template('seleccion.html', error_message=error_message or "", columns=columns,
                           table_name=table_name, selected_columns=selected_columns, records=records)

# Ruta para mostrar el mapa de calor de correlación
@app.route('/mostrar-mapa-de-calor', methods=['POST'])
def mostrar_resultado_correlacion():
    selected_tables = request.form.get('table_names').split(',')
    selected_columns = request.form.getlist('variables[]')

    print("Selected tables:", selected_tables)
    print("Selected columns:", selected_columns)

    # Llamar a la función analizar_correlación del módulo regresión
    result = analizar_correlacion(', '.join(selected_tables), selected_columns)

    if "message" in result:
        # Si el resultado contiene un mensaje, mostrar el mapa de calor
        message = result["message"]
        heatmap_path = result["heatmap_path"]
        return render_template('resultado_correlacion.html', selected_tables=selected_tables,
                               selected_columns=selected_columns, message=message, heatmap_path=heatmap_path)
    else:
        return result  # Devuelve el resultado (puede ser un error o el mapa de calor)

#----------------------------REGRESION----------------------------

# Ruta que permite al usuario seleccionar las variables para el análisis de regresión en una tabla de la base de datos.
@app.route('/seleccionar-variables-regresion/<table_name>', methods=['GET', 'POST'])
def seleccionar_variables_regresion(table_name):
    error_message = None
    selected_columns = []

    # Obtener la lista de columnas y registros de la tabla seleccionada
    columns, records = get_table_info(table_name)

    if request.method == 'POST':
        selected_columns = request.form.getlist('variables')

    # Renderizar la plantilla 'seleccion.html' con los datos obtenidos
    return render_template('seleccion.html', error_message=error_message or "", columns=columns,
                           table_name=table_name, selected_columns=selected_columns, records=records)

# Función para verificar si las variables son válidas para una tabla
def verificar_variables(selected_tables, x_variable, y_variable):
    try:
        # Iterar sobre todas las tablas seleccionadas
        for selected_table in selected_tables:
            # Obtener los nombres de las columnas para la tabla actual
            column_names = obtener_nombres_columnas(selected_table)
            
            # Verificar si las variables son válidas para la tabla actual
            if x_variable not in column_names:
                raise Exception(f"La variable X '{x_variable}' no es válida para la tabla '{selected_table}'.")
            if y_variable not in column_names:
                raise Exception(f"La variable Y '{y_variable}' no es válida para la tabla '{selected_table}'.")
        
        # Si ha pasado la verificación para todas las tablas, devolver True
        return True
    except Exception as e:
        # Si se encuentra alguna excepción, devolver el mensaje de error
        return str(e)

# Ruta para realizar la regresión lineal
@app.route('/regresion_lineal', methods=['POST'])
def regresion_lineal():
    error_message = ""
    success_message = ""
    prediction = None
    beta_0 = None
    beta_1 = None
    r_squared = None
    correlation_coefficient = None
    x_variable = None
    y_variable = None
    registros = None

    if request.method == 'POST':
        selected_tables = request.form.get('table_names').split(',')
        x_variable = request.form.get('x_variable')
        y_variable = request.form.get('y_variable')

        print("Selected tables:", selected_tables)
        print("x_variable:", x_variable)
        print("y_variable:", y_variable)
    
        try:
            # Verificar si las variables son válidas para la tabla seleccionada
            #verification_result = verificar_variables(', '.join(selected_tables), x_variable, y_variable)
            verification_result = all(verificar_variables(selected_table, x_variable, y_variable) for selected_table in selected_tables)

            print(f"registros: {registros}")
            print(f"Valores recibidos antes del if - selected_tables:¨{selected_tables}, x_variable: {x_variable}, y_variable: {y_variable}" )
            
            if verification_result is True:
                
                print(f"Valores recibidos despues del if - selected_tables:¨{selected_tables}, x_variable: {x_variable}, y_variable: {y_variable}" )
               

                # Realizar la regresión lineal
                (
                    success_message, error_message, prediction, beta_0, beta_1,
                    r_squared, correlation_coefficient, registros
                ) = regresion_model.realizar_regresion_lineal(', '.join(selected_tables), x_variable, y_variable)
                print(f"registros: {registros}")
                print(f"verificar los valores obteniendo regresion, beta_0: {beta_0}")
                print(f"verificar los valores obteniendo regresion, beta_1: {beta_1}")
                print(f"verificar los valores obteniendo regresion, r_squared: {r_squared}")
                print(f"verificar los valores obteniendo regresion, correlation_coefficient: {correlation_coefficient}")
                
            else: print(f"la variables NO son válidas para la tabla seleccionada")
            
        except ValueError:
            error_message = "Ingresa un valor válido para X."

    # Crear un contexto con los resultados y renderizar la plantilla 'resultado_regresion.html'
    context = {
        'table_name': ', '.join(selected_tables),
        'x_variable': x_variable,
        'y_variable': y_variable,
        'error_message': error_message,
        'success_message': success_message,
        'prediction': prediction,
        'beta_0': beta_0,
        'beta_1': beta_1,
        'r_squared': r_squared,
        'correlation_coefficient': correlation_coefficient,
        'registros': registros
    }

    print(f"valores antes de pasar al contexto, beta_0: {beta_0}")
    print(f"valores antes de pasar al contexto, beta_1: {beta_1}")
    print(f"valores antes de pasar al contexto, r_squared: {r_squared}")
    print(f"valores antes de pasar al contexto, correlation_coefficient: {correlation_coefficient}")
    
    return render_template('resultado_regresion.html', **context)

#----------------------------PREDICCION----------------------------

# Ruta para realizar la predicción
@app.route('/realizar-prediccion', methods=['POST'])
def realizar_prediccion():
    result = None

    if request.method == 'POST':
        x_variable = request.form.get('x_variable')

        try:
            x_variable = float(x_variable)
            # Realizar la predicción utilizando el modelo de regresión lineal
            result = regresion_model.realizar_prediccion(x_variable)

        except ValueError as ve:
            print(f"Error al convertir a float: {ve}")
            result = {
                'x_variable': None,
                'y_variable': None,
                'prediction': None,
                'error_message': "Ingresa un valor válido para X."
            }

    # Renderizar la plantilla 'resultado_regresion.html' con los resultados de la predicción
    return render_template('resultado_regresion.html', result=result)



#---------------------------- RUTA ETL----------------------------

# Función para conectar a la base de datos MySQL
def connect_to_db():
    try:
        db = get_db_connection()
        if db is not None:
            return db
        else:
            return None
    except mysql.connector.Error as err:
        return None
    
    # Función para obtener la lista de tablas en la base de datos OK
def get_table_list_etl():
    db = connect_to_db()
    if db is not None:
        cursor = db.cursor()
        cursor.execute("SHOW TABLES")
        tables = [table[0] for table in cursor]
        table_list = []
        for table_name in tables:
            columns = get_table_columns_etl(table_name)
            table_list.append({'name': table_name, 'columns': columns})
        db.close()
        return table_list
    else:
        return []
    
# Función para obtener los registros de una tabla
def get_table_records_etl(table_name):
    db = connect_to_db()
    if db is not None:
        cursor = db.cursor()
        cursor.execute(f"SELECT * FROM {table_name}")
        records = cursor.fetchall()
        db.close()
        print(f"Registros obtenidos de la tabla {table_name}: {records}")
        return records
    else:
        return []
    
    
    # Función para obtener las columnas de una tabla y sus registros
def get_table_info_etl(table_name):
    columns = get_table_columns_etl(table_name)
    records = get_table_records_etl(table_name)
    return columns, records

# Función para obtener nombres y tipos de datos de las columnas
def get_table_columns_etl(table_name):
    db = connect_to_db()
    if db is not None:
        cursor = db.cursor()
        cursor.execute(f"SHOW COLUMNS FROM {table_name}")
        columns = [{'name': col[0], 'type': col[1]} for col in cursor.fetchall()]
        db.close()
        return columns
    else:
        return []
    

# Función para obtener los nombres de las columnas de una tabla
def obtener_nombres_columnas_etl(selected_table):
    try:
        db = get_db_connection()
        if db.is_connected():
            cursor = db.cursor()
            query = f"SHOW COLUMNS FROM {selected_table}"
            cursor.execute(query)
            result = cursor.fetchall()
            db.close()

            if result:
                column_names = [row[0] for row in result]
                return column_names
            else:
                raise Exception(f"No se encontraron columnas en la tabla '{selected_table}'.")
        else:
            raise Exception("La conexión a la base de datos no está activa.")
    except Exception as e:
        raise Exception(f"Error al conectar a la base de datos: {e}")


# Ruta para ver y seleccionar datos de una tabla
@app.route('/ver-datos', methods=['GET', 'POST'])
def ver_datos():
    tables = get_table_list_etl()
    table_name = None
    columns = []
    records = []

    if request.method == 'POST':
        table_name = request.form.get('table_name')
        columns, records = get_table_info_etl(table_name)
        
        print(f"Registros obtenidos después de la carga: {records}")

    return render_template('datos.html', tables=tables, table_name=table_name, columns=columns, records=records)

#--------------- EDITAR REGISTRO ---------------
# Ruta para editar un registro
@app.route('/editar/<table_name>/<int:record_id>', methods=['GET', 'POST'])
def editar_registro(table_name, record_id):
    columns = obtener_nombres_columnas_etl(table_name)
    record = obtener_registro_etl(table_name, record_id)

    if request.method == 'POST':
        # Obtener los nuevos valores desde el formulario
        nuevos_valores = {}
        for columna in columns:
            nuevos_valores[columna] = request.form[columna]

        # Actualizar el registro en la base de datos
        actualizar_registro_etl(table_name, record_id, nuevos_valores)

        # Redirigir a la página de ver datos
        return redirect(url_for('ver_datos'))

    return render_template('editar_registro.html', table_name=table_name, columns=columns, record=record)


# Función para obtener un registro por ID
def obtener_registro_etl(table_name, record_id):
    db = connect_to_db()
    if db is not None:
        cursor = db.cursor()
        cursor.execute(f"SELECT * FROM {table_name} WHERE id = %s", (record_id,))
        record = cursor.fetchone()
        db.close()
        return record
    else:
        return None

# Función para actualizar un registro por ID
def actualizar_registro_etl(table_name, record_id, nuevos_valores):
    db = connect_to_db()
    if db is not None:
        cursor = db.cursor()

        # Construir la consulta SQL para la actualización
        update_query = f"UPDATE {table_name} SET "
        update_query += ', '.join([f"{column} = %s" for column in nuevos_valores.keys()])
        update_query += f" WHERE id = %s"

        # Crear una tupla con los nuevos valores y el ID del registro
        valores_actualizados = tuple(nuevos_valores.values()) + (record_id,)

        # Ejecutar la consulta
        cursor.execute(update_query, valores_actualizados)

        # Confirmar la transacción y cerrar la conexión
        db.commit()
        db.close()

#--------------- ELIMINAR REGISTRO ---------------

# Ruta para eliminar un registro
@app.route('/eliminar/<table_name>/<int:record_id>', methods=['GET', 'POST'])
def eliminar_registro(table_name, record_id):
    if request.method == 'POST':
        eliminar_registro_etl(table_name, record_id)

        # Redirigir a la página de ver datos después de eliminar
        return redirect(url_for('ver_datos'))

    return render_template('eliminar_registro.html', table_name=table_name, record_id=record_id)


# Función para obtener y luego eliminar un registro por ID
def eliminar_registro_etl(table_name, record_id):
    db = connect_to_db()
    if db is not None:
        cursor = db.cursor()

        # Obtener los datos del registro antes de eliminarlo
        cursor.execute(f"SELECT * FROM {table_name} WHERE id = %s", (record_id,))
        record = cursor.fetchone()

        if record:
            print(f"Datos del registro antes de eliminar: {record}")

            # Eliminar el registro de la base de datos solo si se encontraron datos
            cursor.execute(f"DELETE FROM {table_name} WHERE id = %s", (record_id,))
            db.commit()

            print("Registro eliminado correctamente.")
        else:
            print("No se encontraron datos para el registro.")

        db.close()

        return record  # Devolvemos los datos del registro eliminado o None si no se encontraron datos

#--------------- DESCARGAR PLANTILLA EXCEL ---------------

# Función para obtener nombres y tipos de datos de las columnas
def get_table_columns_etl_des(table_name):
    db = connect_to_db()
    if db is not None:
        cursor = db.cursor()
        cursor.execute(f"SHOW COLUMNS FROM {table_name}")
        columns = cursor.fetchall()
        db.close()
        return columns
    else:
        return []

# Ruta para descargar la plantilla
@app.route('/descargar-plantilla/<table_name>', methods=['GET'])
def descargar_plantilla(table_name):
    # Obtener información sobre las columnas y tipos de datos de la tabla desde la base de datos
    columnas_tipos = get_table_columns_etl_des(table_name)
    columnas = [col[0] for col in columnas_tipos]
    tipos_de_datos = [col[1] for col in columnas_tipos]

    # Crear una plantilla de Pandas DataFrame con columnas y tipos de datos
    df_template = pd.DataFrame(columns=columnas)

    # Crear un libro de trabajo de Excel y una hoja
    wb = Workbook()
    ws = wb.active

    # Crear estilos para cada tipo de dato
    styles_dict = {
        'int': styles.NamedStyle(name='int_style', number_format=styles.numbers.FORMAT_NUMBER),
        'float': styles.NamedStyle(name='float_style', number_format=styles.numbers.FORMAT_NUMBER_00),
        'str': styles.NamedStyle(name='str_style')
    }

    # Agregar encabezados a la hoja de cálculo con el estilo correspondiente
    for col_num, (columna, tipo_de_dato) in enumerate(zip(columnas, tipos_de_datos), 1):
        cell = ws.cell(row=1, column=col_num, value=columna)
        estilo_name = f'{tipo_de_dato}_style'

        # Verificar si el estilo ya existe antes de agregarlo
        if estilo_name not in wb.named_styles:
            estilo = styles_dict.get(tipo_de_dato, styles.NamedStyle(name=estilo_name))
            wb.add_named_style(estilo)

        # Aplicar el estilo a la celda
        cell.style = estilo_name

    # Crear un archivo temporal para almacenar el libro de trabajo
    temp_file, temp_filename = tempfile.mkstemp(suffix='.xlsx')
    os.close(temp_file)
    
    # Guardar el libro de trabajo en el archivo temporal
    wb.save(temp_filename)

    # Devolver el archivo para su descarga
    return send_file(
        temp_filename,
        download_name=f'{table_name}_template.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


#--------------- CARGAR DATOS DESDE EXCEL ---------------

# Ruta para cargar registros desde un archivo de Excel
@app.route('/cargar-registros/<table_name>', methods=['POST'])
def cargar_registros(table_name):
    if request.method == 'POST':
        try:
            # Obtener el archivo enviado desde el formulario
            archivo_excel = request.files['archivo_excel']

            if archivo_excel.filename == '':
                flash('Error: No se seleccionó ningún archivo.', 'error')
                return redirect(url_for('ver_datos'))

            # Validar que el archivo sea un archivo Excel
            if not archivo_excel.filename.endswith('.xlsx'):
                flash('Error: El archivo debe tener la extensión .xlsx.', 'error')
                return redirect(url_for('ver_datos'))

            # Cargar el archivo Excel en un DataFrame de Pandas
            df = pd.read_excel(archivo_excel)

            # Ajustes para manejar tipos de datos y valores nulos de manera dinámica
            for columna in df.columns:
                if pd.api.types.is_numeric_dtype(df[columna]):
                    # Convertir a Python int si la columna es de tipo numérico
                    df[columna] = pd.to_numeric(df[columna], errors='coerce', downcast='integer')
                elif pd.api.types.is_datetime64_any_dtype(df[columna]):
                    # Convertir a fecha si la columna es de tipo fecha
                    formato_fecha = "%Y-%m-%d"  # Ajusta el formato según tus fechas
                    df[columna] = pd.to_datetime(df[columna], errors='coerce', format=formato_fecha)

            # Llenar los valores nulos con 0 o cadena vacía, ajusta según sea necesario
            df.fillna(0, inplace=True)

                        # Asegurarse de que las fechas sean reconocidas correctamente y no sean numpy.datetime64
            # Asegurarse de que las fechas sean reconocidas correctamente y no sean numpy.datetime64
            for columna_fecha in df.select_dtypes(include=['datetime64']).columns:
                formato_fecha = "%Y-%m-%d"  # Ajusta el formato según tus fechas
                df[columna_fecha] = pd.to_datetime(df[columna_fecha], errors='coerce').dt.date


            print("DataFrame desde el archivo Excel:")
            print(df)

            # Conectar a la base de datos
            db = connect_to_db()
            if db is not None:
                # Obtener los nombres de las columnas de la tabla
                columnas_tabla = obtener_nombres_columnas_etl(table_name)
                print("Columnas en la tabla de la base de datos:")
                print(columnas_tabla)

                # Filtrar solo las columnas que existen en la base de datos
                columnas_validas = [col for col in df.columns if col in columnas_tabla]

                # Filtrar el DataFrame para incluir solo las columnas válidas
                df = df[columnas_validas]

                # Verificar si hay alguna columna requerida que falta en el archivo Excel
                columnas_faltantes = [col for col in columnas_tabla if col not in columnas_validas]
                if columnas_faltantes:
                    flash(f'Error: Las siguientes columnas son requeridas y faltan en el archivo Excel: {", ".join(columnas_faltantes)}', 'error')
                    return redirect(url_for('ver_datos'))

                # Verificar duplicados en la clave primaria antes de la inserción
                clave_primaria = obtener_clave_primaria(table_name)
                if clave_primaria is not None:
                    registros_duplicados = df[df.duplicated(subset=clave_primaria, keep=False)]
                    num_registros_duplicados = registros_duplicados.shape[0]
                    if num_registros_duplicados > 0:
                        mensaje_duplicados = f'Error: Se encontraron registros duplicados en la clave primaria: {registros_duplicados.to_dict(orient="records")}'
                        flash(mensaje_duplicados, 'error')
                        return redirect(url_for('ver_datos'))

                # Convertir el DataFrame a una lista de tuplas para su inserción en la base de datos
                # Asegúrate de que las fechas sean objetos datetime de Python y no numpy.datetime64
                registros = [tuple(str(x) if isinstance(x, pd.Timestamp) else int(x) if isinstance(x, (np.integer, np.int64, np.int8, np.int16, np.int32)) else x for x in row) for row in df.to_records(index=False)]

                print("Registros a ser insertados en la base de datos:")
                print(registros)

                # Construir la consulta SQL para la inserción
                consulta_insercion = f"INSERT INTO {table_name} ({', '.join(columnas_validas)}) VALUES ({', '.join(['%s' for _ in columnas_validas])}) ON DUPLICATE KEY UPDATE id=id"

                print("Consulta de inserción:")
                print(consulta_insercion)

                # Ejecutar la consulta de inserción
                with db.cursor() as cursor:
                    cursor.executemany(consulta_insercion, registros)

                # Confirmar la transacción y cerrar la conexión
                db.commit()
                db.close()

                # Imprimir el número de filas afectadas
                num_filas_afectadas = cursor.rowcount
                print(f"Número de filas afectadas: {num_filas_afectadas}")

                # Imprimir el número de registros duplicados
                print(f"Número de registros duplicados: {num_registros_duplicados}")

                # Imprimir el número de registros nulos
                num_registros_nulos = df.isnull().sum().sum()
                print(f"Número de registros nulos: {num_registros_nulos}")

                flash(f'Registros cargados exitosamente. Filas afectadas: {num_filas_afectadas}, Duplicados: {num_registros_duplicados}, Nulos: {num_registros_nulos}', 'success')
            else:
                flash('Error al conectar a la base de datos.', 'error')
        except Exception as e:
            flash(f'Error al cargar registros: {str(e)}', 'error')
            print(f'Error al cargar registros: {str(e)}')

        # Redirigir a la página de ver datos
        return redirect(url_for('ver_datos'))


# Esta función devolver la clave primaria de la tabla dada
def obtener_clave_primaria(table_name):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        cursor.execute(f"DESCRIBE {table_name}")
        columns = cursor.fetchall()
        db.close()

        # Filtrar las columnas que son clave primaria
        primary_key_columns = [col[0] for col in columns if "PRI" in col[3]]

        return primary_key_columns
    except Exception as e:
        print(f"Error al obtener clave primaria: {str(e)}")
        return None




#--------------- DASHBOARD ---------------

@app.route('/dashboard', methods=['GET'])
def mostrar_dashboard():
    # Obtener los datos de la tabla "clientes"
    db = connect_to_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM clientes")
    clientes_data = cursor.fetchall()

    #___________OBTENER DATOS___________
    
    # Obtener los datos de la tabla "productos"
    cursor.execute("SELECT * FROM productos")
    productos_data = cursor.fetchall()

    # Obtener los datos de la tabla "lotes"
    cursor.execute("SELECT * FROM lotes")
    lotes_data = cursor.fetchall()

    # Obtener los datos de la tabla intermedia "ProductosLotes"
    cursor.execute("SELECT * FROM ProductosLotes")
    productos_lotes_data = cursor.fetchall()

    # Obtener los datos de la tabla "ventasproductos" con información adicional de productos
    cursor.execute("""
        SELECT vp.id, vp.Precio_Lote, vp.Unidades_Vendidas_Producto, vp.Subtotal_Producto,
            vp.id_Venta, vp.id_Producto, p.Nombre_Producto
        FROM ventasproductos vp
        INNER JOIN productos p ON vp.id_Producto = p.id
    """)
    ventasproductos_data = cursor.fetchall()

    # Obtener los datos de la tabla "ventas"
    cursor.execute("SELECT * FROM ventas")
    ventas_data = cursor.fetchall()
    
    cursor.execute("SELECT * FROM insumos")
    insumos_data = cursor.fetchall()
    

    db.close()

    #___________CREAR DATA FRAME___________
    
    # clientes
    df_clientes = pd.DataFrame(clientes_data, columns=["id", "Nombre", "Celular", "Tipo_Documento", "Num_Documento", "Ciudad", "Direccion", "Tipo_Cliente"])

    # productos
    df_productos = pd.DataFrame(productos_data, columns=["id", "Nombre_Producto"])

    # lotes
    df_lotes = pd.DataFrame(lotes_data, columns=["id", "Fecha_Fabricacion", "Fecha_Vencimiento", "Dias_Caducar",
                                                "Unidades_Preparadas", "Costo_Unitario", "Precio_Detal", "Precio_Mayorista",
                                                "Unidades_Totales_vendidas", "Ventas_Totales", "Unidades_Disponibles", "Utilidad_Total"])

    # ventasproductos y productos
    df_ventasproductos = pd.DataFrame(ventasproductos_data, columns=["id", "Precio_Lote", "Unidades_Vendidas_Producto", "Subtotal_Producto", "id_Venta", "id_Producto", "Nombre_Producto"])

    # ProductosLotes
    df_productos_lotes = pd.DataFrame(productos_lotes_data, columns=["id", "id_Producto", "id_Lote"])

    # Fusionar las tablas "productos" y "lotes" utilizando la tabla intermedia "ProductosLotes"
    df_productos_lotes = pd.merge(df_productos_lotes, df_productos, left_on='id_Producto', right_on='id')
    df_productos_lotes = pd.merge(df_productos_lotes, df_lotes, left_on='id_Lote', right_on='id')

    #ventas
    df_ventas = pd.DataFrame(ventas_data, columns=["id", "Tipo_Venta", "Subtotal_Venta", "Fecha_venta", "Descuento", "Costo_Envio", "Total_Venta"])

    # Crear DataFrame
    df_insumos = pd.DataFrame(insumos_data, columns=["id", "Nombre", "Fecha_Compra", "Fecha_Vencimiento", "Dias_Caducar",
                                                "Precio_Total", "Cantidad_Adquirida", "Cantidad_Disponible", "Cantidad_Usada",
                                                "Unidad_Medida", "Desperdicio", "Proveedor", "Precio_Unitario"])

    # Convertir las columnas de fecha a tipo datetime
    df_insumos['Fecha_Compra'] = pd.to_datetime(df_insumos['Fecha_Compra'])
    df_insumos['Fecha_Vencimiento'] = pd.to_datetime(df_insumos['Fecha_Vencimiento'])
    
    # Ordenar DataFrame por fecha de compra
    df_insumos = df_insumos.sort_values(by='Fecha_Compra')
    
    #___________CACULOS___________
    
    # Contar el número de clientes por ciudad
    conteo_ciudad = df_clientes['Ciudad'].value_counts()

    # Calcular el Total_Venta por Tipo_Venta
    total_por_tipo_venta = df_ventas.groupby('Tipo_Venta')['Total_Venta'].sum().reset_index()
    
    

    #________________GRAFICO lotes asociados a un producto y sus Unidades_Preparadas ___________________
    
    # Crear un gráfico de barras para visualizar los lotes asociados a un producto y sus Unidades_Preparadas
    fig_productos_lotes = px.bar(df_productos_lotes, x='id_Lote', y='Unidades_Preparadas',
                                color='Nombre_Producto', labels={'Unidades_Preparadas': 'Unidades Preparadas'},
                                title='Unidades Preparadas por Lote y Producto',
                                hover_data=['Nombre_Producto', 'Unidades_Preparadas']) 

    # Configurar el diseño del gráfico
    fig_productos_lotes.update_layout(
        xaxis_title='ID del Lote',
        yaxis_title='Unidades Preparadas',
        font=dict(family='Arial', size=14, color='black'),
        title_font=dict(size=20, family='Arial'),
    )

    # Convertir el gráfico a HTML
    graph_html_productos_lotes = fig_productos_lotes.to_html(full_html=False)

        
    
    
    #________________GRAFICO cantidades insumos ___________________
    
    # Crear gráfico de líneas para las cantidades de insumos
    fig_insumos = px.line(df_insumos, x='Nombre', y=['Cantidad_Adquirida', 'Cantidad_Disponible', 'Cantidad_Usada'],
                        labels={'value': 'Cantidad', 'variable': 'Tipo'},
                        title='Cantidades de Insumos')

    # Configurar el diseño del gráfico de insumos
    fig_insumos.update_layout(
        xaxis_title='Insumos',
        yaxis_title='Cantidad',
        font=dict(family='Arial', size=14, color='black'),
        title_font=dict(size=20, family='Arial'),
        height=600,
    )

    # Convertir el gráfico a HTML
    graph_html_insumos = fig_insumos.to_html(full_html=False)


    #________________GRAFICO clientes____________________

    # Crear un gráfico de barras personalizado para clientes
    fig_clientes = px.bar(
        x=conteo_ciudad.index,
        y=conteo_ciudad.values,
        labels={'x': 'Ciudad', 'y': 'Número de Clientes'},
        title=f'Número de Clientes por Ciudad (Total Clientes: {len(df_clientes)})',
        color=conteo_ciudad.values,
        color_continuous_scale='Viridis',
        text=conteo_ciudad.values,
        height=500,
    )

    # Configurar el diseño del gráfico de clientes
    fig_clientes.update_layout(
        showlegend=False,
        xaxis_title='Ciudad',
        yaxis_title='Número de Clientes',
        font=dict(family='Arial', size=14, color='black'),
        title_font=dict(size=20, family='Arial'),
    )

    # Convertir el gráfico de clientes a HTML
    graph_html_clientes = fig_clientes.to_html(full_html=False)

    #________________GRAFICO productos Unidades disponibles ____________________

    # Crear un gráfico de barras para productos
    fig_productos = px.bar(df_productos_lotes, x='Nombre_Producto', y='Unidades_Disponibles', title='Unidades Disponibles por Producto')

    # Configurar el diseño del gráfico de productos
    fig_productos.update_layout(
        xaxis_title='Productos Vascu SAS',
        yaxis_title='Unidades Disponibles',
        font=dict(family='Arial', size=14, color='black'),
        title_font=dict(size=20, family='Arial'),
    )

    # Convertir el gráfico de productos a HTML
    graph_html_productos = fig_productos.to_html(full_html=False)

    #________________GRAFICO ventasproductos ____________________

    # Crear un gráfico de barras apiladas para visualizar las ventas de productos
    fig_ventasproductos = px.bar(
        df_ventasproductos,
        x='id_Venta',  # Cada transacción en el eje X
        y='Subtotal_Producto',  # Altura de la barra según el subtotal del producto
        color='Nombre_Producto',  # Color de la barra según el nombre del producto
        labels={'id_Venta': 'ID de Venta', 'Subtotal_Producto': 'Subtotal del Producto'},
        title='Ventas de Productos por Transacción',
    )
    # Configurar el diseño del gráfico de ventas de productos
    fig_ventasproductos.update_layout(
        xaxis_title='ID de Venta',
        yaxis_title='Subtotal del Producto',
        font=dict(family='Arial', size=14, color='black'),
        title_font=dict(size=20, family='Arial'),
        barmode='stack',  # Barras apiladas
    )
    # Convertir el gráfico de ventas de productos a HTML
    graph_html_ventasproductos = fig_ventasproductos.to_html(full_html=False)

    #________________GRAFICO Total_Venta por Tipo_Venta  ____________________

    # Crear un gráfico de torta para visualizar el Total_Venta por Tipo_Venta
    fig_ventas_tipo = px.pie(
        total_por_tipo_venta,
        names='Tipo_Venta',
        values='Total_Venta',
        title='Total de Ventas por Tipo de Venta',
        hole=0.3,  # Agujero en el centro para hacerlo un donut chart
    )
    # Convertir el gráfico de ventas por tipo a HTML
    graph_html_ventas_tipo = fig_ventas_tipo.to_html(full_html=False)



    return render_template('dashboard.html', graph_html_clientes=graph_html_clientes, graph_html_productos=graph_html_productos,
                           graph_html_ventasproductos=graph_html_ventasproductos, graph_html_ventas_tipo=graph_html_ventas_tipo,
                           graph_html_insumos=graph_html_insumos, graph_html_productos_lotes=graph_html_productos_lotes)



#--------------- INICIO APP ---------------

# Iniciar la aplicación si este archivo se ejecuta directamente
if __name__ == '__main__':
    app.run(debug=True)
